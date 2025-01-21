from decimal import Decimal
from email.policy import default

import openpyxl
from django.core.exceptions import ValidationError
from django import forms
from zaek.models import ZaekPrice, ClassificationPriceProduct


def validate_columns_df(list_columns, df, sheet_name=None):
    header = None
    try:
        list_columns_set = set(list_columns)
        workbook = openpyxl.load_workbook(df, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationError(f'Лист с именем "{sheet_name}" не найден в файле.')
            workbook = workbook[sheet_name]

        for i, row in enumerate(workbook.iter_rows(values_only=True)):
            if list_columns_set.issubset(set(row)):
                header = i
                break
        if header is None:
            raise ValidationError(f'Нет нужных колонок {", ".join(list_columns)}')
        return header

    except Exception as e:
        raise ValidationError(f'Ошибка при проверки файла {e}')




class BaseExcelForm(forms.Form):
    file = forms.FileField(
        label='Файл',
        required=True
    )
     # Допустимые форматы файлов

    def clean_file(self):
        format_file = self.__class__.format_file
        columns = self.__class__.columns
        sheet_name = self.__class__.sheet_name

        file = self.cleaned_data['file']
        if not file:
            raise ValidationError('Вы ничего не загрузили')

        if not str(file).endswith(format_file):
            raise ValidationError(f'Формат должен быть {format_file}')

        if columns:
            validate_columns_df(columns,file,sheet_name)


        return file


class PriceLoadZaekForm(BaseExcelForm):
    format_file = '.xlsx'
    columns = None
    sheet_name = None


class PriceLoadZaekGroupsForm(BaseExcelForm):
    format_file = '.csv'
    columns = None
    sheet_name = None


class ConsolidatedTableForm(BaseExcelForm):
    format_file = '.xlsx'
    columns = ['Арт','Кол']
    sheet_name = 'Запрос'


class FindObjectsExcelForm(BaseExcelForm):
    format_file = '.xlsx'
    columns = ['Арт','Кол']
    sheet_name = 'Запрос'

class PreparingDataForLoadingForm(BaseExcelForm):
    format_file = '.xlsx'
    columns = ['Арт','Кол']
    sheet_name = 'Запрос'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['сводная'] = forms.BooleanField(
            label='Сводная',
            required=False,
            initial=False
        )



        # Добавляем динамические поля для каждого объекта из модели ClassificationPriceProduct
        # for product in ClassificationPriceProduct.objects.all():
        #     field_name = f'price_{product.id}'  # создаем имя поля
        #     self.fields[field_name] = forms.DecimalField(
        #         label=product.name,
        #         required=False,
        #         decimal_places=2,
        #         max_digits=10,
        #         min_value=0,
        #         initial=0
        #     )

    # def clean(self):
    #     cleaned_data = super().clean()
    #     products = ClassificationPriceProduct.objects.all()
    #
    #     for product in products:
    #         field_name = f'price_{product.id}'
    #         price = cleaned_data.get(field_name)
    #
    #         if price is None:
    #             raise ValidationError(f'Поле {product.name} не должно быть пустым.')
    #
    #         self._validate_price(product.name, price)
    #
    #     return cleaned_data
    #
    # def _validate_price(self, name, price):
    #
    #     if not isinstance(price, (int, float, Decimal)):
    #         print(type(price))
    #         raise ValidationError(f'Скидка для {name} должна быть числом.')
    #     if price < 0:
    #         raise ValidationError(f'Скидка для {name} не может быть отрицательной.')
    #     if price > 10000:
    #         raise ValidationError(f'Скидка {name} не может быть больше 10 000.')

