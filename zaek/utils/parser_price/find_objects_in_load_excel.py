import math
from decimal import Decimal
from io import BytesIO
from django.db.models import Prefetch
from openpyxl.styles import PatternFill, Alignment
from zaek.consts_zaek import split_parent_base, split_parent_obj
from zaek.models import Product, ClientClassificationDiscount, VolumeAtDiscount, ClassificationPriceProduct, VolumeLevel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle
import pandas as pd


def art_format(art):
    try:
        return str(int(art))
    except:
        return str(art)


class CreateDictFindObjectsExcel:
    ruble_style = NamedStyle(name="RubleStyle")
    ruble_style.number_format = '#,##0.00 "₽"'

    def __init__(
            self,
            request_products=None,
            request_products_dict=None,
            df = None
    ):
        self.request_products = request_products
        self.request_products_dict = request_products_dict
        self.df=df
        self.test_header = {}
        self.wb = Workbook()
        self.ws = self.wb.active
        self.start_header = 2
        self.ROC = 48
        self.classification_list = set()
        self.summ_packaging_norm_result_all=0
        self.volume_level = None
        self.volume_level_discounts ={}
        self.color_warning = 'cc3333'
        self.color_true ="00a86b"
        self.dict_costumer_num_row = {}
        self.end_row = 1
        self.base_wight_col = 20
        self.wraptext = True


    # def get_price_group_list(self,price_group_list):
    #     """Добавляем уровни родителей в объект"""
    #     for group in price_group_list.split(split_parent_base):
    #         level_value = group.split(split_parent_obj)
    #         level = int(level_value[0])
    #         value = level_value[1]
    #         level_name = f'Уровень: {level}'
    #         if level_name not in self.price_group_list_columns_to_sort:
    #             self.price_group_list_columns_to_sort.append(level_name)
    #         self.product[f'Уровень: {level}'] = value

    def get_float(self, obj):
        try:
            print(f"DEBUG: raw obj={obj}, type={type(obj)}")

            if isinstance(obj, Decimal):
                if obj.is_nan():
                    raise ValueError("Значение Decimal является NaN")
                float_obj = float(obj)

            else:
                float_obj = float(obj)
                if math.isnan(float_obj):
                    raise ValueError("Значение float является NaN")

            if float_obj < 0:
                ret = 0.0
            else:
                ret = float_obj

        except (ValueError, TypeError) as e:

            ret = 0.0
        return ret

    def get_or_create_header(self,key):
        """Создаю колонки по запросу"""
        obj = self.test_header.get(key)
        if obj:
            return obj
        col_letter =get_column_letter(len(self.test_header)+1)
        self.test_header[key] = col_letter

        self.ws[f"{col_letter}{self.start_header - 1}"] = key
        self.ws.column_dimensions[col_letter].width = self.base_wight_col
        self.ws[f"{col_letter}{self.start_header - 1}"].alignment = Alignment(wrapText=True)
        return col_letter


    def color_section(self,color_type_client_type):
        """Настройки закраски клеток"""
        return PatternFill(start_color=color_type_client_type, end_color=color_type_client_type, fill_type="solid")

    def obj_money_style_rub(self,obj):
        """Денежный стиль в рублях"""
        obj.number_format = CreateDictFindObjectsExcel.ruble_style.number_format

    def sellable_quantity(self, pack_size, quantity):
        """Вычисление кратности упаковки"""
        try:
            if quantity % pack_size == 0:
                return quantity
            else:
                return ((quantity // pack_size) + 1) * pack_size
        except Exception as e:
            print(f"Ошибка в sellable_quantity: pack_size={pack_size}, quantity={quantity}, ошибка={e}")
            return 0

    def fide_volume_level_project_discount(self):
        """Получение скидки за счет объема (Проектная скидка)"""
        try:

            value = self.summ_packaging_norm_result_all
            print(value,'ss')
            discounts = VolumeLevel.objects.all().order_by('start_value')


            if value < discounts.first().start_value:
                return None

            for i, disc in enumerate(discounts):
                if value < disc.start_value:

                    return discounts[i-1]

            return discounts.last()
        except:
            return None

    def create_volume_level_discounts(self):
        """Создание словаря проектных скидок """
        volume_discounts = self.volume_level.discounts.all()
        for v in volume_discounts:
            if v.type_client.type not in self.volume_level_discounts:
                self.volume_level_discounts[v.type_client.type] = {}
            self.volume_level_discounts[v.type_client.type][v.classification_price_product.name] = v.discount

    def cut_name_row(self,value):
        num_row = self.dict_costumer_num_row.get(value)
        if not num_row:
            self.end_row+=1
            self.dict_costumer_num_row[value] = self.end_row
            return self.end_row
        return num_row

    def get_project_discount(self,cli,classif,discount,color):
        """Получение по ключу проектную скидку """
        try:
            return self.volume_level_discounts[cli][classif], color
        except:
            return discount,self.color_warning

    def get_previous_column_letter(self,col_letter):
        """Возвращает предыдущую букву колонки."""
        col_index = column_index_from_string(col_letter)
        if col_index > 1:
            previous_col_index = col_index - 1
            return get_column_letter(previous_col_index)
        else:
            raise ValueError("Нет предыдущей колонки для первой колонки (A).")

    def align_text_center(self):
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal="center", vertical="center")


    def create_summ_formula_and_title(self,letter,value,color=None):
        ssn = 'Сумма с НДС'
        sbn = 'Сумма без НДС'
        ssn_row = self.cut_name_row(ssn)
        sbn_line = self.cut_name_row(sbn)

        self.ws[f'{letter}{ssn_row}'] = value
        self.ws[f'{self.get_previous_column_letter(letter)}{ssn_row}'] = ssn
        self.ws[f'{letter}{sbn_line}'] = f'={letter}{ssn_row} * 0.8'
        self.ws[f'{self.get_previous_column_letter(letter)}{sbn_line}'] = sbn

        if color:
            self.ws[f'{letter}{ssn_row}'].fill = self.color_section(color)
            self.ws[f'{letter}{sbn_line}'].fill = self.color_section(color)
            self.ws[f'{self.get_previous_column_letter(letter)}{ssn_row}'].fill = self.color_section(color)
            self.ws[f'{self.get_previous_column_letter(letter)}{sbn_line}'].fill = self.color_section(color)

        self.obj_money_style_rub(self.ws[f'{letter}{ssn_row}'])
        self.obj_money_style_rub(self.ws[f'{letter}{sbn_line}'])


    def create_dict(self):
        print('create_dict')
        """Формируем объект в виде словаря"""
        list_request_products_dict_keys= list(self.request_products_dict.keys())
        count_request_products = len(list_request_products_dict_keys)
        dict_request_products = {obj.art: obj for obj in self.request_products}
        dict_type_client = {}





        col_letter_art = self.get_or_create_header('Артикул')
        col_letter_classification = self.get_or_create_header('Номенклатура')
        col_letter_kol = self.get_or_create_header('Кол-во')
        col_letter_price_with_nds = self.get_or_create_header('Цена (с НДС) руб.')
        col_letter_price_not_nds = self.get_or_create_header('Цена (без НДС) руб.')
        col_letter_summ_base = self.get_or_create_header('Сумма по прайсу с НДС')
        col_letter_summ_base_not_nds = self.get_or_create_header('Сумма по прайсу с без НДС')
        col_letter_packaging_norm = self.get_or_create_header('Норма упаковки')
        col_letter_packaging_norm_result = self.get_or_create_header('Количество с округлением норма упаковки')
        col_letter_summ_packaging_norm_result = self.get_or_create_header('Сумма с округлением норма упаковки c НДС')
        col_letter_summ_packaging_norm_result_not_nds = self.get_or_create_header('Сумма с округлением норма упаковки c без НДС')

        for i, row in self.df.iterrows():
            idx = i + self.start_header

            art_obj = art_format(row['Арт'])
            count_odj = row['Кол']
            request_product = dict_request_products.get(art_obj)

            if request_product:

                self.ws[f'{col_letter_art}{idx}'] = request_product.art

                self.ws[f'{col_letter_classification}{idx}'] = request_product.name

                count_product = Decimal(self.get_float(count_odj))
                self.ws[f'{col_letter_kol}{idx}'] = count_product

                self.ws[f'{col_letter_classification}{idx}'] = request_product.classification.name
                self.classification_list.add(request_product.classification.name)

                self.ws[f'{col_letter_price_with_nds}{idx}'] = request_product.price_with_nds
                self.obj_money_style_rub(self.ws[f'{col_letter_price_with_nds}{idx}'])

                self.ws[f'{col_letter_summ_base}{idx}'] = \
                    f"={col_letter_price_with_nds}{idx} * {col_letter_kol}{idx}"
                self.obj_money_style_rub(self.ws[f'{col_letter_summ_base}{idx}'])


                self.ws[f'{col_letter_price_not_nds}{idx}'] = request_product.price_not_nds
                self.obj_money_style_rub(self.ws[f'{col_letter_price_not_nds}{idx}'])

                self.ws[f'{col_letter_summ_base_not_nds}{idx}'] = \
                    f"={col_letter_price_not_nds}{idx} * {col_letter_kol}{idx}"
                self.obj_money_style_rub(self.ws[f'{col_letter_summ_base_not_nds}{idx}'])



                request_product_packaging_norm = self.get_float(request_product.packaging_norm)

                self.ws[f'{col_letter_packaging_norm}{idx}'] = request_product_packaging_norm

                self.ws[f'{col_letter_packaging_norm_result}{idx}'] = \
                    f"=IF(MOD({col_letter_kol}{idx}, {col_letter_packaging_norm}{idx}) = 0, " \
                    f"{col_letter_kol}{idx}, " \
                    f"(INT({col_letter_kol}{idx} / {col_letter_packaging_norm}{idx}) + 1) * {col_letter_packaging_norm}{idx})"


                self.ws[f'{col_letter_packaging_norm_result}{idx}'].fill = self.color_section(
                    self.color_warning if count_product % Decimal(request_product_packaging_norm) != 0 else self.color_true
                )

                self.ws[f'{col_letter_summ_packaging_norm_result}{idx}'] = \
                    f"={col_letter_price_with_nds}{idx} * {col_letter_packaging_norm_result}{idx}"

                self.obj_money_style_rub(self.ws[f'{col_letter_summ_packaging_norm_result}{idx}'])

                self.ws[f'{col_letter_summ_packaging_norm_result_not_nds}{idx}'] = \
                    f"={col_letter_price_not_nds}{idx} * {col_letter_packaging_norm_result}{idx}"

                self.obj_money_style_rub(self.ws[f'{col_letter_summ_packaging_norm_result_not_nds}{idx}'])



                sum_obj = self.sellable_quantity(Decimal(request_product_packaging_norm),count_product) * request_product.price_with_nds
                self.summ_packaging_norm_result_all += sum_obj


            else:
                self.ws[f'{col_letter_art}{idx}'] = art_obj



        start_row = self.start_header   # Начальная строка данных
        end_row = self.start_header + count_request_products - 1 # Конечная строка данных
        self.end_row  = self.start_header + count_request_products + 1


        self.volume_level = self.fide_volume_level_project_discount()
        project_discount_flag = True if self.volume_level else False
        if project_discount_flag:
            self.create_volume_level_discounts()


        print("Обозначим проектную скидку")
        """Обозначим проектную скидку"""
        project_discount_flag = True if self.volume_level else False
        project_discount_title = 'Градация за объем проектной скидки от:'
        project_discount_title_row = self.cut_name_row(project_discount_title)
        self.ws[f'A{project_discount_title_row}'] = project_discount_title


        self.ws[f'B{project_discount_title_row}']= (
            self.volume_level.start_value if project_discount_flag else "Не подходит ни под один параметр"
        )

        self.ws[f'B{project_discount_title_row}'].fill = self.color_section(
            self.color_true if project_discount_flag else self.color_warning
        )
        self.obj_money_style_rub(self.ws[f'B{project_discount_title_row}'])

        """Базовый расчет суммы"""
        formula = f"=SUM({col_letter_summ_base}{start_row}:{col_letter_summ_base}{end_row})"
        self.create_summ_formula_and_title(
            letter=col_letter_summ_base,
            value=formula
        )

        formula = f"=SUM({col_letter_summ_packaging_norm_result}{start_row}:{col_letter_summ_packaging_norm_result}{end_row})"
        self.create_summ_formula_and_title(
            letter=col_letter_summ_packaging_norm_result,
            value=formula
        )






        print("Создаем расчет по типам клиентов")
        """Создаем расчет по типам клиентов"""
        for i, row in self.df.iterrows():
            idx = i + self.start_header

            art_obj = art_format(row['Арт'])
            request_product = dict_request_products.get(art_obj)

            if request_product:
                print(request_product)
                for discount in request_product.classification.classification_discounts.all():
                    discount_type_client_type = discount.type_client.type
                    color_type_client_type = discount.type_client.color

                    dict_type_client[discount_type_client_type]=color_type_client_type


                    col_letter_type_client_discount = self.get_or_create_header(f'{discount_type_client_type} скидка')
                    col_letter_type_client_sale_solo = self.get_or_create_header(f'{discount_type_client_type} шт с НДС')
                    col_letter_type_client_sale_summ = self.get_or_create_header(f'{discount_type_client_type} сумма с НДС')


                    """Гарантированные скидки  скидки от объема """
                    self.ws[f'{col_letter_type_client_discount}{idx}'] = discount.discount
                    self.ws[f'{col_letter_type_client_discount}{idx}'].fill = self.color_section(color_type_client_type)
                    self.ws[f'{col_letter_type_client_sale_solo}{idx}'] = (
                        f'={col_letter_price_with_nds}{idx}*(1-{col_letter_type_client_discount}{idx}/100)'
                    )
                    self.obj_money_style_rub(self.ws[f'{col_letter_type_client_sale_solo}{idx}'])
                    self.ws[f'{col_letter_type_client_sale_solo}{idx}'].fill = self.color_section(color_type_client_type)

                    self.ws[f'{col_letter_type_client_sale_summ}{idx}'] = (
                        f'={col_letter_packaging_norm_result}{idx}*{col_letter_type_client_sale_solo}{idx}'
                    )
                    self.obj_money_style_rub(self.ws[f'{col_letter_type_client_sale_summ}{idx}'])
                    self.ws[f'{col_letter_type_client_sale_summ}{idx}'].fill = self.color_section(color_type_client_type)

                    """Проектные скидки от объема """
                    col_letter_projects_type_client_discount = self.get_or_create_header(f'{discount_type_client_type} проектная скидка')
                    col_letter_projects_type_client_sale_solo = self.get_or_create_header(f'{discount_type_client_type} проектная шт с НДС')
                    col_letter_projects_type_client_sale_summ = self.get_or_create_header(f'{discount_type_client_type} проектная сумма кратно уп. с НДС')
                    col_letter_projects_type_client_sale_solo_not_nds = self.get_or_create_header(f'{discount_type_client_type} проектная шт без НДС')
                    col_letter_projects_type_client_sale_summ_not_nds = self.get_or_create_header(f'{discount_type_client_type} проектная сумма кратно уп. без НДС')


                    project_discount,color_project_discount = self.get_project_discount(
                        discount_type_client_type,
                        request_product.classification.name,
                        discount.discount,
                        color_type_client_type
                    )

                    self.ws[f'{col_letter_projects_type_client_discount}{idx}'] = project_discount
                    self.ws[f'{col_letter_projects_type_client_discount}{idx}'].fill = self.color_section(color_project_discount)

                    self.ws[f'{col_letter_projects_type_client_sale_solo}{idx}'] = (
                        f'={col_letter_price_with_nds}{idx}*(1-{col_letter_projects_type_client_discount}{idx}/100)'
                    )
                    self.obj_money_style_rub(self.ws[f'{col_letter_projects_type_client_sale_solo}{idx}'])
                    self.ws[f'{col_letter_projects_type_client_sale_solo}{idx}'].fill = self.color_section(color_type_client_type)

                    self.ws[f'{col_letter_projects_type_client_sale_summ}{idx}'] = (
                        f'={col_letter_packaging_norm_result}{idx}*{col_letter_projects_type_client_sale_solo}{idx}'
                    )
                    self.obj_money_style_rub(self.ws[f'{col_letter_projects_type_client_sale_summ}{idx}'])
                    self.ws[f'{col_letter_projects_type_client_sale_summ}{idx}'].fill = self.color_section(color_type_client_type)





                    self.ws[f'{col_letter_projects_type_client_sale_solo_not_nds}{idx}'] = (
                        f'={col_letter_price_not_nds}{idx}*(1-{col_letter_projects_type_client_discount}{idx}/100)'
                    )

                    self.obj_money_style_rub(self.ws[f'{col_letter_projects_type_client_sale_solo_not_nds}{idx}'])
                    self.ws[f'{col_letter_projects_type_client_sale_solo_not_nds}{idx}'].fill = self.color_section(
                        color_type_client_type)


                    self.ws[f'{col_letter_projects_type_client_sale_summ_not_nds}{idx}'] = (
                        f'={col_letter_packaging_norm_result}{idx}*{col_letter_projects_type_client_sale_solo_not_nds}{idx}'
                    )
                    self.obj_money_style_rub(self.ws[f'{col_letter_projects_type_client_sale_summ_not_nds}{idx}'])
                    self.ws[f'{col_letter_projects_type_client_sale_summ_not_nds}{idx}'].fill = self.color_section(
                        color_type_client_type)





        print("Расчет сумм по типам клиентов")
        """Расчет сумм по типам клиентов"""
        for type_client in list(dict_type_client.keys()):
            col_letter_client = self.get_or_create_header(f'{type_client} сумма с НДС')
            formula = f"=SUM({col_letter_client}{start_row}:{col_letter_client}{end_row})"
            self.create_summ_formula_and_title(
                letter=col_letter_client,
                value=formula,
                color=dict_type_client.get(type_client)
            )

            col_letter_client = self.get_or_create_header(f'{type_client} проектная сумма кратно уп. с НДС')
            formula = f"=SUM({col_letter_client}{start_row}:{col_letter_client}{end_row})"
            self.create_summ_formula_and_title(
                letter=col_letter_client,
                value=formula,
                color=dict_type_client.get(type_client)
            )

        self.align_text_center()
        print('ww')
        return self.wb



def find_objects_in_load_excel(file=None):
        pd.read_excel(file)
        data_dict = {}
        for i , row in pd.read_excel(file).iterrows():
            r_art=art_format(row['Арт'])
            art = str(r_art)
            data_dict[art] = row['Кол']

        data_keys_list = list(data_dict.keys())

        products = Product.objects.prefetch_related(
            'classification',
                Prefetch(
                    'classification__classification_discounts',  # Скидки, связанные через ClassificationPriceProduct
                    queryset=ClientClassificationDiscount.objects.select_related(
                        'type_client',
                        'classification_price_product',
                    )
                ),

            Prefetch(
                'classification__classification_volume_discounts',  # Скидки за объем, связанные через ClassificationPriceProduct
                queryset=VolumeAtDiscount.objects.select_related('volume_level', 'type_client'),
            )
        ).filter(art__in=data_keys_list)

        if products:

            new_excel_file = CreateDictFindObjectsExcel
            excel_creator = new_excel_file(
                request_products = products,
                request_products_dict = data_dict,
                df = pd.read_excel(file)
            ).create_dict()

            output = BytesIO()
            excel_creator.save(output)
            output.seek(0)
            return output

