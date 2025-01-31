import hashlib
import math
from decimal import Decimal
from io import BytesIO
from openpyxl.styles import PatternFill, NamedStyle, Alignment

from zaek.models import Product, ClassificationPriceProduct
import pandas as pd
from openpyxl import Workbook
from zaek.utils.parser_price.consolidated_table import create_dict_consolidated_table
from openpyxl.utils import get_column_letter, column_index_from_string


class PreDataLoadingInto:
    ruble_style = NamedStyle(name="RubleStyle")
    ruble_style.number_format = '#,##0.00 "₽"'

    def __init__(
            self,
            excel ,
            slodnaya_bool
    ):
        self.close_df = excel
        self.slodnaya_bool = slodnaya_bool
        self.classification_pp = [obj.name for obj in ClassificationPriceProduct.objects.all()]
        self.wb = Workbook()
        self.ws = self.wb.active
        self.df_open = None
        self.list_object = []
        self.header_row = 1
        self.last_column = 0
        self.dict_classification_pp = {}
        self.dict_header = {}
        self.color_classification_pp= {}
        self.base_wight_col = 20
        self.wraptext = True
        self.let_row_summ_nds_ny = ''
        self.let_row_summ_no_nds_ny = ''
        self.let_row_summ_nds = ''
        self.let_row_summ_no_nds = ''
        self.color_warning = 'cc3333'
        self.color_true ="00a86b"


    def obj_money_style_rub(self,obj):
        """Денежный стиль в рублях"""
        obj.number_format = PreDataLoadingInto.ruble_style.number_format

    def pars_data_loading_into(self):
        self.df_open = pd.read_excel(self.close_df)
        self.list_object = set()
        if self.slodnaya_bool:
            cons = create_dict_consolidated_table(self.df_open)
            self.df_open = pd.DataFrame(list(cons.items()), columns=['Арт', 'Кол'])
            self.list_object = (list(cons.keys()))
        else:
            for i , row in self.df_open.iterrows():
               self.list_object.add(self.art_format(row['Арт']))

        self.create_dict_classification_pp_and_write_in_df()
        self.iter_data()
        return self.wb



    def generate_color_from_number(self,number):
        value_str = str(number)
        value_hash = hashlib.md5(value_str.encode()).hexdigest()
        r = int(value_hash[:2], 16)
        g = int(value_hash[2:4], 16)
        b = int(value_hash[4:6], 16)
        return f"{r:02X}{g:02X}{b:02X}"



    def get_letter_column(self,key):
        """Готовим хедар для таблицы """
        wight_col = self.base_wight_col
        value = self.dict_header.get(key)
        if value:
            return value
        col_letter = get_column_letter(len(self.dict_header)+1)
        self.dict_header[key]= col_letter

        self.ws[f"{col_letter}{self.header_row}"] = key
        len_key = len(key)
        if len_key < self.base_wight_col:
            wight_col = len_key +2

        self.ws.column_dimensions[col_letter].width = wight_col
        # self.ws[f"{col_letter}{self.header_row }"].alignment = Alignment(wrapText=True)
        self.ws[f"{col_letter}{self.header_row }"].fill = self.color_section('98c793')
        return col_letter

    def color_section(self,color_type_client_type):
        """Настройки закраски клеток"""
        return PatternFill(start_color=color_type_client_type, end_color=color_type_client_type, fill_type="solid")

    def create_dict_classification_pp_and_write_in_df(self):
        """Создаем дефолтные значения"""
        for idx , value in enumerate(self.classification_pp,1):
            self.ws[f"A{idx}"] = value
            self.dict_classification_pp[value] = f'B{idx}'
            self.ws[f"B{idx}"] = 0
            self.header_row=idx
            color = self.generate_color_from_number(idx)
            self.color_classification_pp[value] = color
            self.ws[f"A{idx}"].fill = self.color_section(color)
            self.ws[f"B{idx}"].fill = self.color_section(color)

        self.header_row += 1
        self.ws[f"A{self.header_row}"] = 'Конкурент'
        self.ws[f"B{self.header_row}"] = 'Chint'
        self.dict_classification_pp['Конкурент'] = f'B{self.header_row }'

        self.header_row += 1
        self.ws[f"A{self.header_row}"] = 'Сумма без НДС'
        self.let_row_summ_no_nds = f"B{self.header_row}"
        self.header_row += 1
        self.ws[f"A{self.header_row}"] = 'Сумма без c НДС'
        self.let_row_summ_nds = f"B{self.header_row}"

        self.header_row += 1
        self.ws[f"A{self.header_row}"] = 'Н/У Сумма без НДС'
        self.let_row_summ_no_nds_ny = f"B{self.header_row}"
        self.header_row += 1
        self.ws[f"A{self.header_row}"] = 'Н/У Сумма без c НДС'
        self.let_row_summ_nds_ny = f"B{self.header_row}"

        self.header_row+=3

    def get_float(self, obj):
        try:
            if isinstance(obj, Decimal):
                if obj.is_nan():
                    raise ValueError("Значение Decimal является NaN")
                float_obj = float(obj)

            else:
                float_obj = float(obj)
                if math.isnan(float_obj):
                    raise ValueError("Значение float является NaN")

            if float_obj < 0:
                ret = 0.0
            else:
                ret = float_obj

        except (ValueError, TypeError) as e:

            ret = 0.0
        return ret


    def art_format(self,art):
        try:
            return str(int(art))
        except:
            return str(art)


    def iter_data(self):
        """Обрабатываем данные из таблицы """
        products_filter = Product.objects.select_related('classification').filter(art__in=self.list_object)
        dict_products_filter = {obj.art: obj for obj in products_filter}
        row_num = 0
        letter_column_art = self.get_letter_column('Арт')
        letter_column_col = self.get_letter_column('Кол')
        letter_column_sale = self.get_letter_column('Скидка')
        letter_column_competitor =self.get_letter_column('Конкурент')
        letter_column_classification= self.get_letter_column('Классификация ПП')
        letter_column_name = self.get_letter_column('Наименование')
        letter_column_price_no_nds = self.get_letter_column('Цена (без НДС) руб.')
        letter_column_price_nds = self.get_letter_column('Цена (с НДС) руб.')
        letter_column_price_no_nds_sale = self.get_letter_column('Скидка Цена (без НДС) руб.')
        letter_column_price_nds_sale = self.get_letter_column('Скидка Цена (с НДС) руб.')
        letter_column_price_no_nds_summ_sale = self.get_letter_column('Скидка Сумма (без НДС) руб. ')
        letter_column_price_nds_summ_sale = self.get_letter_column('Скидка Сумма (с НДС) руб.')
        letter_packaging_norm = self.get_letter_column('Норма упаковки')
        letter_packaging_norm_result = self.get_letter_column('Результат Норма упаковки')
        letter_column_price_no_nds_summ_sale_ny = self.get_letter_column('Н/У Скидка Сумма (без НДС) руб. ')
        letter_column_price_nds_summ_sale_ny = self.get_letter_column('Н/У Скидка Сумма (с НДС) руб.')

        for i, row in self.df_open.iterrows():
            row_num = self.header_row + 1 + i
            art = self.art_format(row['Арт'])
            kol = self.get_float(row['Кол'])
            obj = dict_products_filter.get(art)

            if obj:
                self.ws[f"{letter_column_name}{row_num}"] = obj.name
                self.ws[f"{letter_column_art}{row_num}"] = obj.art
                self.ws[f"{letter_column_col}{row_num}"] = kol

                classification_letter_num = self.dict_classification_pp.get(
                    f'{obj.classification.name}',
                    'Нет данных'
                )

                self.ws[f"{letter_column_sale}{row_num}"] = f"={classification_letter_num}"
                comp_letter_num = self.dict_classification_pp.get('Конкурент', 'Нет данных')
                self.ws[f"{letter_column_competitor}{row_num}"] = f"={comp_letter_num}"
                self.ws[f"{letter_column_classification}{row_num}"] = obj.classification.name

                classification_color = self.color_classification_pp.get(obj.classification.name)
                if classification_color:
                    self.ws[f"{letter_column_classification}{row_num}"].fill = self.color_section(classification_color)
                    self.ws[f"{letter_column_sale}{row_num}"].fill = self.color_section(classification_color)


                self.ws[f"{letter_packaging_norm}{row_num}"] = obj.packaging_norm

                self.ws[f'{letter_packaging_norm_result}{row_num}'] = \
                    f"=IF(MOD({letter_column_col}{row_num}, {letter_packaging_norm}{row_num}) = 0, " \
                    f"{letter_column_col}{row_num}, " \
                    f"(INT({letter_column_col}{row_num} / {letter_packaging_norm}{row_num}) + 1) * {letter_packaging_norm}{row_num})"

                self.ws[f'{letter_packaging_norm_result}{row_num}'].fill = self.color_section(
                    self.color_warning if kol % obj.packaging_norm != 0 else self.color_true
                )



                """Цена за шт """
                self.ws[f"{letter_column_price_no_nds}{row_num}"] = obj.price_not_nds
                self.obj_money_style_rub(self.ws[f"{letter_column_price_no_nds}{row_num}"])
                self.ws[f"{letter_column_price_nds}{row_num}"] = obj.price_with_nds
                self.obj_money_style_rub( self.ws[f"{letter_column_price_nds}{row_num}"])

                """Цна Шт со скидкой"""
                self.ws[f'{letter_column_price_no_nds_sale}{row_num}'] = (
                    f'={letter_column_price_no_nds}{row_num}*(1-{letter_column_sale}{row_num}/100)'
                )
                self.obj_money_style_rub(self.ws[f'{letter_column_price_no_nds_sale}{row_num}'])

                self.ws[f'{letter_column_price_nds_sale}{row_num}'] = (
                    f'={letter_column_price_nds}{row_num}*(1-{letter_column_sale}{row_num}/100)'
                )
                self.obj_money_style_rub(self.ws[f'{letter_column_price_nds_sale}{row_num}'])

                """Расчет сумм"""
                self.ws[f"{letter_column_price_no_nds_summ_sale}{row_num}"] = \
                    f'={letter_column_col}{row_num}*{letter_column_price_no_nds_sale}{row_num}'
                self.obj_money_style_rub(self.ws[f"{letter_column_price_no_nds_summ_sale}{row_num}"])

                self.ws[f"{letter_column_price_nds_summ_sale}{row_num}"] = \
                    f'={letter_column_col}{row_num}*{letter_column_price_nds_sale}{row_num}'
                self.obj_money_style_rub(self.ws[f"{letter_column_price_nds_summ_sale}{row_num}"])

                self.ws[f"{letter_column_price_no_nds_summ_sale_ny}{row_num}"] = \
                    f'={letter_packaging_norm_result}{row_num}*{letter_column_price_no_nds_sale}{row_num}'
                self.obj_money_style_rub(self.ws[f"{letter_column_price_no_nds_summ_sale_ny}{row_num}"])

                self.ws[f"{letter_column_price_nds_summ_sale_ny}{row_num}"] = \
                     f'={letter_packaging_norm_result}{row_num}*{letter_column_price_nds_sale}{row_num}'
                self.obj_money_style_rub(self.ws[f"{letter_column_price_nds_summ_sale_ny}{row_num}"])

            else:
                self.ws[f"{letter_column_art}{row_num}"] = art
                self.ws[f"{letter_column_col}{row_num}"] = kol
                self.ws[f"{letter_column_name}{row_num}"] = "Не найден"


        """Итого по суммам по колонкам"""
        self.ws[self.let_row_summ_no_nds] = \
            f"=SUM({letter_column_price_no_nds_summ_sale}{self.header_row +1 }:{letter_column_price_no_nds_summ_sale}{row_num})"
        self.obj_money_style_rub(self.ws[f"{letter_column_price_no_nds_summ_sale}{row_num + 1}"])
        self.obj_money_style_rub(self.ws[self.let_row_summ_no_nds])

        self.ws[self.let_row_summ_nds] = \
            f"=SUM({letter_column_price_nds_summ_sale}{self.header_row + 1}:{letter_column_price_nds_summ_sale}{row_num})"
        self.obj_money_style_rub(self.ws[f"{letter_column_price_nds_summ_sale}{row_num + 1}"])
        self.obj_money_style_rub(self.ws[self.let_row_summ_nds])

        self.ws[self.let_row_summ_no_nds_ny] = \
            f"=SUM({letter_column_price_no_nds_summ_sale_ny}{self.header_row +1 }:{letter_column_price_no_nds_summ_sale_ny}{row_num})"
        self.obj_money_style_rub(self.ws[f"{letter_column_price_no_nds_summ_sale_ny}{row_num + 1}"])
        self.obj_money_style_rub(self.ws[self.let_row_summ_no_nds_ny])

        self.ws[self.let_row_summ_nds_ny] = \
            f"=SUM({letter_column_price_nds_summ_sale_ny}{self.header_row + 1}:{letter_column_price_nds_summ_sale_ny}{row_num})"
        self.obj_money_style_rub(self.ws[f"{letter_column_price_nds_summ_sale_ny}{row_num + 1}"])
        self.obj_money_style_rub(self.ws[self.let_row_summ_nds_ny])





def preparing_data_loading_into(excel,slodnaya_bool):
    cla = PreDataLoadingInto
    data = cla(excel = excel,slodnaya_bool = slodnaya_bool).pars_data_loading_into()
    output = BytesIO()
    data.save(output)
    output.seek(0)
    return output