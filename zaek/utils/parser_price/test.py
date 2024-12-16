from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Данные
data = [
    {'Артикул': '344592', 'Цена (без НДС) руб.': 6700.00, 'СЩО  3 - х Золотой': 0.04, 'Пенис': 0.04},
    {'Артикул': '285971', 'Цена (без НДС) руб.': 384.00, 'СЩО  3 - х Золотой': 0.04, 'Пенис': 0.04},
]

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Данные"

# Заголовки
headers = list(data[0].keys())
headers += ['СЩО  3 - х Золотой - Цена (без НДС) руб.', 'Пенис - Цена (без НДС) руб.']
ws.append(headers)

# Заполнение данных и добавление формул
for row_idx, item in enumerate(data, start=2):  # Нумерация строк с 2, так как 1 строка - заголовки
    ws[f"A{row_idx}"] = item['Артикул']
    ws[f"B{row_idx}"] = item['Цена (без НДС) руб.']
    ws[f"C{row_idx}"] = item['СЩО  3 - х Золотой']
    ws[f"D{row_idx}"] = item['Пенис']

    # Формулы для новых колонок
    ws[f"E{row_idx}"] = f"=B{row_idx} * (1 - C{row_idx})"  # СЩО 3 - х Золотой
    ws[f"F{row_idx}"] = f"=B{row_idx} * (1 - D{row_idx})"  # Пенис

# Устанавливаем ширину колонок для удобства
for col_idx, header in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 25

# Сохраняем файл
wb.save("/home/dima/Python/django/ckd_web/zaek/utils/parser_price/v.xlsx")
print("Файл успешно создан: dynamic_columns.xlsx")