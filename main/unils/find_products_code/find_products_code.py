import os
from io import BytesIO

from core_app.redis_cli import RedisClientMain
from main.forms import validate_columns_df
from main.models import *
import openpyxl
import pandas as pd
from main.unils.parser_magadel.const import parent_split
from main.unils.find_products_code.const import sheet_name, const_columns, new_columns_dict
from openpyxl.styles import PatternFill



def adjust_column_width(worksheet):
    """
    Настраивает ширину столбцов по ширине самого длинного текста с небольшим запасом.
    """
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2  # Добавляем небольшой запас
        worksheet.column_dimensions[column].width = adjusted_width


def open_df(new_path):
    header = validate_columns_df(
        list(const_columns.values()),
        new_path,
        sheet_name
    )
    excel_data_df = pd.read_excel(new_path, header=header, sheet_name=sheet_name, engine='openpyxl')
    import numpy as np
    excel_data_df = excel_data_df.replace({np.nan: None})
    return excel_data_df, header


def create_df_columns(df, column):
    """
    Создаем колонку если ее нет
    """

    if isinstance(column, (list, tuple)):
        for col in column:
            if col not in df.columns:
                df[col] = None

    elif isinstance(column, str):
        if column not in df.columns:
            df[column] = None


def get_level_groups(level_group_columns_name, group, df, i):
    """
    Добавляем в level_group_columns_name название колонок
    И проверим на наличие колонок в дф
    """
    for index, value in enumerate(group.split(parent_split)):
        column = f'Уровень: {index}'
        if column not in level_group_columns_name:
            level_group_columns_name.append(column)
            create_df_columns(df, column)
        df.at[i, column] = value


def get_float(x):
    try:

        return float(x)
    except:
        return 0

def sum_price_product(price,quantity,sum_product):
    try:

        q= price * quantity
        sum_product += q
        return q,sum_product
    except:
        return 0

def update_values_object(object, df, index_row, quantity,sum_product,maga_name):
    free_balance = object.get('free_balance', 0)
    sum_possible_deliveries = object.get('sum_possible_deliveries', 0)
    list_possible_deliveries = object.get('list_possible_deliveries', 'Нет значения')
    unit = object.get('unit', 'Нет значения')
    request_quantity = get_float(quantity)
    sum_possible_deliveries_free_balance = (sum_possible_deliveries + free_balance) - request_quantity
    name = object.get('name', 'Нет значения')
    price = object.get('price',0)

    df.at[index_row,  new_columns_dict['Свободный остаток']] = free_balance
    df.at[index_row,  new_columns_dict['Свободный остаток - Кол']] = free_balance - request_quantity
    df.at[index_row,  new_columns_dict['Свободные поступления']] = sum_possible_deliveries
    df.at[index_row,  new_columns_dict['Свободный остаток - Свободные поступления - Кол']] = sum_possible_deliveries_free_balance
    df.at[index_row,  new_columns_dict['Информация о поступлениях']] = '\n'.join(list_possible_deliveries.split(','))
    df.at[index_row,  new_columns_dict['Описание']] = name
    df.at[index_row,  new_columns_dict['ЕдИзм']] = unit
    df.at[index_row,  new_columns_dict['Прайс без НДС']] =price
    sp,sum_product = sum_price_product(price,quantity,sum_product)
    df.at[index_row,  new_columns_dict['Итого без НДС']] = sp
    df.at[index_row, new_columns_dict['V']] = maga_name
     

    return sum_product

def create_new_def(df,file):
    output_file = file
    dict_prod = {}
    for index_row, row in df.iterrows():
        code = row[const_columns['Код']]
        if code:
            count_code = get_float(row[const_columns['Кол']])
            if code in dict_prod:
                dict_prod[code]['Кол'] += count_code
            else:
                dict_prod[code] = {
                    const_columns['Код']: code,
                    const_columns['Кол']: count_code
                }

    list_prod = list(dict_prod.values())
    df = pd.DataFrame(list_prod)

    # Запись DataFrame в указанный лист Excel-файла
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    return df

def get_maga_name(rm):
    try:
        maga_name = rm.get_maga(Magadel)

        return maga_name['name']
    except:
        return 'Нет данных'


def find_products(file):
    df, header = open_df(file)
    rm = RedisClientMain()
    products, groups = rm.get_products_parents(model_product=ProductDKCMagadel, model_group=GroupProductDKCMagadel)

    maga_name = get_maga_name(rm)
    df = create_new_def(df,file)


    base_columns = (list(new_columns_dict.values()))

    create_df_columns(df, base_columns)

    level_group_columns_name = []
    last_index = 0
    sum_product=0






    for index_row, row in df.iterrows():
        last_index=index_row
        product_code = row.get('Код')


        if product_code:
            product = products.get(product_code)

            if product:

                quantity = get_float(row.get('Кол'))
                sum_product = update_values_object(product, df, index_row, quantity,sum_product,maga_name)
                group = product.get('parent')
                if group:
                    get_level_groups(level_group_columns_name, group, df, index_row)
        else:
            df.at[index_row, 'Описание'] = "Не найдено"

    df.at[last_index + 1, new_columns_dict['Итого без НДС']] = sum_product

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = workbook[sheet_name]

        fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        list_index_to_write_color=[
            df.columns.get_loc(new_columns_dict['Свободный остаток - Кол']) + 1,
            df.columns.get_loc(new_columns_dict['Свободный остаток - Свободные поступления - Кол']) + 1,
        ]

        for row in range(2, worksheet.max_row + 1):
            for inx in list_index_to_write_color:
                cell1 = worksheet.cell(row=row, column=inx)
                if cell1.value:
                    if cell1.value >= 0:
                        cell1.fill = fill_green
                    else:
                        cell1.fill = fill_red


        adjust_column_width(worksheet)
    output.seek(0)

    return output, file.name


def find_products_code_func(file):
    output, file.name = find_products(file)

    return output, file.name
