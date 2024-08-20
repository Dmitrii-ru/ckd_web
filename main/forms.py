from django import forms
from django.core.exceptions import ValidationError
import openpyxl


def validate_columns_df(list_columns, df, sheet_name=None):
    header = None
    try:
        list_columns_set = set(list_columns)
        workbook = openpyxl.load_workbook(df, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationError(f'Лист с именем "{sheet_name}" не найден в файле.')
            workbook = workbook[sheet_name]

        for i, row in enumerate(workbook.iter_rows(values_only=True)):
            if list_columns_set.issubset(set(row)):
                header = i
                break

        if header is None:
            raise ValidationError(f'Нет нужных колонок {list_columns}')
        return header

    except Exception as e:
        raise ValidationError(f'Ошибка при проверки файла {e}')


class FileUploadMagaForm(forms.Form):
    file = forms.FileField()

    def clean_file(self):
        from .unils.parser_magadel.const import const, sheet_name
        file = self.cleaned_data['file']
        if not file:
            raise ValidationError('Вы ничего не загрузили')

        if not str(file).endswith('.xlsx'):
            raise ValidationError('Формат должен быть .xlsx')

        validate_columns_df(
            list(const.values()),
            file,
            sheet_name
        )
        return file


class FileUploadFindProductsCode(forms.Form):
    file = forms.FileField()

    def clean_file(self):
        from .unils.find_products_code.const import const_columns, sheet_name
        file = self.cleaned_data['file']
        if not file:
            raise ValidationError('Вы ничего не загрузили')

        if not str(file).endswith('.xlsx'):
            raise ValidationError('Формат должен быть .xlsx')

        validate_columns_df(
            list(const_columns.values()),
            file,
            sheet_name
        )
        return file
