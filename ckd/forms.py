from django import forms
from django.core.exceptions import ValidationError
import openpyxl

from ckd.models import FileMaga


def validate_columns_df(list_columns, df, sheet_name=None):
    header = None

    list_columns_set = set(list_columns)
    try:
        workbook = openpyxl.load_workbook(df, read_only=True, data_only=True)
    except Exception as e:
        raise ValidationError(f'Не могу открыть файл: {e}')
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f'Лист с именем "{sheet_name}" не найден в файле.')
        workbook = workbook[sheet_name]

    for i, row in enumerate(workbook.iter_rows(values_only=True)):
        if list_columns_set.issubset(set(row)):

            header = i
            break

    if header is None:
        raise ValidationError(f'Нет нужных колонок {list_columns}')
    return header







class FileUploadMagaForm(forms.ModelForm):
    class Meta:
        model = FileMaga
        fields = ['file']

    def clean_file(self):
        from .unils.parser_magadel.const import const, sheet_name
        file = self.cleaned_data['file']
        if not file:
            raise ValidationError('Вы ничего не загрузили')

        if not str(file).endswith('.xlsx'):
            raise ValidationError('Формат должен быть .xlsx')


        validate_columns_df(
            list(const.values()),
            file,
            sheet_name
        )
        return file


class FileUploadFindProductsCode(forms.Form):
    file = forms.FileField()

    def clean_file(self):
        from .unils.find_products_code.const import const_columns, sheet_name
        file = self.cleaned_data['file']
        if not file:
            raise ValidationError('Вы ничего не загрузили')

        if not str(file).endswith('.xlsx'):
            raise ValidationError('Формат должен быть .xlsx')

        validate_columns_df(
            list(const_columns.values()),
            file,
            sheet_name
        )
        return file
